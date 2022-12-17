import csv
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
import matplotlib.pyplot as plt
import matplotlib
import numpy as np

title = 0
salaryMin = 1
salaryMax = 2
curr = 3
area = 4
timePublic = 5

currency = {
    "AZN": 35.68,
    "BYR": 23.91,
    "EUR": 59.90,
    "GEL": 21.74,
    "KGS": 0.76,
    "KZT": 0.13,
    "RUR": 1,
    "UAH": 1.64,
    "USD": 60.66,
    "UZS": 0.0055,
}

class Report:
    def __init__(self, filename, name):
        self.filename = filename
        self.name = name
        self.years = list(range(2007, 2023))
        self.yearsSums = {}
        self.yearsLength = {}
        self.yearsSumsCurr = {}
        self.yearsLengthCurr = {}
        self.cities = []
        self.citiesSums = {}
        self.citiesLength = {}
        self.vacanciesLength = 0
        self.ansCitiesSums = {}
        self.citiesPartitions = {}
        self.read_file()
        self.calculate_file()
        self.Wb = Workbook()

    def read_file(self):
        kol = False
        with open(self.filename, encoding="utf-8") as file:
            reader = csv.reader(file)
            for data in reader:
                if not kol:
                    kol = True
                    title = data.index("name")
                    salaryMin = data.index("salary_from")
                    salaryMax = data.index("salary_to")
                    curr = data.index("salary_currency")
                    area = data.index("area_name")
                    timePublic = data.index("published_at")
                else:
                    row = data.copy()
                    if all(row):
                        currencyYear = int(data[timePublic].split("-")[0])
                        currencySalary = (int(float(data[salaryMax])) + int(float(data[salaryMin]))) * currency[data[curr]] // 2
                        currencyName = data[title]
                        currencyCity = data[area]
                        self.yearsSums[currencyYear] = self.yearsSums.get(currencyYear, 0) + currencySalary
                        self.yearsLength[currencyYear] = self.yearsLength.get(currencyYear, 0) + 1
                        if profession in currencyName:
                            self.yearsSumsCurr[currencyYear] = self.yearsSumsCurr.get(currencyYear, 0) + currencySalary
                            self.yearsLengthCurr[currencyYear] = self.yearsLengthCurr.get(currencyYear, 0) + 1
                        if currencyCity not in self.cities:
                            self.cities.append(currencyCity)
                        self.citiesSums[currencyCity] = self.citiesSums.get(currencyCity, 0) + currencySalary
                        self.citiesLength[currencyCity] = self.citiesLength.get(currencyCity, 0) + 1
                        self.vacanciesLength += 1

    def calculate_file(self):
        for i in self.years:
            if self.yearsSums.get(i, None):
                self.yearsSums[i] = int(self.yearsSums[i] // self.yearsLength[i])
            if self.yearsSumsCurr.get(i, None):
                self.yearsSumsCurr[i] = int(self.yearsSumsCurr[i] // self.yearsLengthCurr[i])

        for j in self.cities:
            self.citiesSums[j] = int(self.citiesSums[j] // self.citiesLength[j])
        interestingCities = [city for city in self.cities if self.citiesLength[city] >= self.vacanciesLength // 100]
        self.ansCitiesSums = {key: self.citiesSums[key] for key in sorted(interestingCities, key=lambda x: self.citiesSums[x], reverse=True)[:10]}
        self.citiesPartitions = {key: float("{:.4f}".format(self.citiesLength[key] / self.vacanciesLength)) for key in sorted(interestingCities, key=lambda x: self.citiesLength[x] / self.vacanciesLength, reverse=True)[:10]}

    def print_file(self):
        print("Динамика уровня зарплат по годам:", self.yearsSums)
        print("Динамика количества вакансий по годам:", self.yearsLength)
        if not len(self.yearsSumsCurr):
            self.yearsSumsCurr[2022] = 0
        print("Динамика уровня зарплат по годам для выбранной профессии:", self.yearsSumsCurr)
        if not len(self.yearsLengthCurr):
            self.yearsLengthCurr[2022] = 0
        print("Динамика количества вакансий по годам для выбранной профессии:", self.yearsLengthCurr)
        print("Уровень зарплат по городам (в порядке убывания):", self.ansCitiesSums)
        print("Доля вакансий по городам (в порядке убывания):", self.citiesPartitions)

    def generate_excel(self):
        self.yearsStatSheet = self.Wb.create_sheet(title="Статистика по годам")
        self.citiesStatSheet = self.Wb.create_sheet(title="Статистика по городам")
        self.Wb.remove(self.Wb["Sheet"])
        sd = Side(border_style='thin', color="000000")
        self.border = Border(right=sd, top=sd, bottom=sd, left=sd)
        self.headerAlignment = Alignment(horizontal='left')
        self.dataAlignment = Alignment(horizontal='right')
        self.citiesStatSheet["a1"] = 12
        self.reportYears()
        self.reportCities()
        self.fit_cells()
        self.Wb.save('report.xlsx')

    def reportYears(self):
        headers = ["Год", "Средняя зарплата", "Средняя зарплата - " + self.name,
                   "Количество вакансий", "Количество вакансий - " + self.name]
        self.setHeaders(self.yearsStatSheet, headers)

        matrix = []
        for row in range(len(self.yearsSums)):
            key = list(self.yearsSums.keys())[row]
            appendable = [key, self.yearsSums[key], self.yearsSumsCurr[key], self.yearsLength[key],
                          self.yearsLengthCurr[key]]
            matrix.append(appendable)

        self.fillMatrix(self.yearsStatSheet, matrix, offset=(0, 1))

    def fillMatrix(self, sheet, matrix, offset=(0, 0)):
        for row in range(len(matrix)):
            for col in range(len(matrix[0])):
                address = f"{get_column_letter(col + 1 + offset[0])}{row + 1 + offset[1]}"
                sheet[address] = matrix[row][col]
                sheet[address].border = self.border
                sheet[address].alignment = self.dataAlignment
                sheet.column_dimensions[get_column_letter(col + 1)].auto_size = 1

    def setHeaders(self, sheet, headers, offset=(0, 0)):
        for col in range(0, len(headers)):
            address = f"{get_column_letter(col + 1 + offset[0])}{1 + offset[1]}"
            sheet[address] = headers[col]
            sheet[address].border = self.border
            sheet[address].alignment = self.headerAlignment
            sheet[address].font = Font(bold=True)
            sheet.column_dimensions[get_column_letter(col + 1)].auto_size = 1

    def fit_cells(self):
        for sheetName in self.Wb.sheetnames:
            sheet = self.Wb[sheetName]
            for col in range(1, sheet.max_column + 1):
                width = None
                for row in range(1, sheet.max_row + 1):
                    value = sheet[f"{get_column_letter(col)}{row}"].value
                    if value is not None and (width is None or len(str(value)) > width):
                        width = len(str(value))
                if width is not None:
                    sheet.column_dimensions[f"{get_column_letter(col)}"].width = width + 2
                else:
                    sheet.column_dimensions[f"{get_column_letter(col)}"].width = + 2

    def reportCities(self):
        headersPayment = ["Город", "Уровень зарплат"]
        headersPercent = ["Город", "Доля вакансий"]
        self.setHeaders(self.citiesStatSheet, headersPayment)
        self.setHeaders(self.citiesStatSheet, headersPercent, (3, 0))

        self.dataAlignment = Alignment(horizontal='left')
        self.fillMatrix(self.citiesStatSheet, [[i] for i in self.ansCitiesSums.keys()], offset=(0, 1))
        matrix = {key: f"{(val * 10000) // 1 / 100}%" for key, val in self.citiesPartitions.items()}
        self.fillMatrix(self.citiesStatSheet, [[i] for i in list(matrix.keys())], offset=(3, 1))
        self.dataAlignment = Alignment(horizontal='right')
        self.fillMatrix(self.citiesStatSheet, [[i] for i in list(self.ansCitiesSums.values())], offset=(1, 1))
        self.fillMatrix(self.citiesStatSheet, [[i] for i in list(matrix.values())], offset=(4, 1))

    def generate_image(self):
        matplotlib.rc("font", size=8)
        fig, ((ax1, ax2), (ax3, ax4)) = plt.subplots(nrows=2, ncols=2)
        width = 0.3
        x = np.arange(len(self.yearsSums.keys()))
        payment1 = ax1.bar(x - width / 2, self.yearsSums.values(), width, label="средняя з/п")
        payment2 = ax1.bar(x + width / 2, self.yearsSumsCurr.values(), width, label=f"з/п {self.name}")

        ax1.grid(True, axis="y")
        ax1.set_title("Уровень зарплат по годам")
        ax1.set_xticks(np.arange(len(self.yearsSums.keys())), self.yearsSums.keys(), rotation=90)
        ax1.bar_label(payment1, fmt="")
        ax1.bar_label(payment2, fmt="")
        ax1.legend(prop={"size": 6})

        ax2.grid(True, axis="y")
        ax2.set_title("Количество вакансий по годам")
        x = np.arange(len(self.yearsSums.keys()))
        ax2.set_xticks(x, self.yearsSums.keys(), rotation=90)
        vac1 = ax2.bar(x - width / 2, self.yearsSums.values(), width, label="Количество вакансий")
        vac2 = ax2.bar(x + width / 2, self.yearsSumsCurr.values(), width, label=f"Количество вакансий\n{self.name}")
        ax2.bar_label(vac1, fmt="")
        ax2.bar_label(vac2, fmt="")
        ax2.legend(prop={"size": 6})

        ax3.grid(True, axis="x")
        y = np.arange(len(list(self.ansCitiesSums.keys())))
        ax3.set_yticks(y, map(lambda s: s.replace(" ", "\n").replace("-", "\n"), self.ansCitiesSums.keys()))
        ax3.invert_yaxis()
        ax3.barh(y, self.ansCitiesSums.values())
        ax3.set_title("Уровень зарплат по городам")

        ax4.set_title("Доля вакансий по городам")
        other = 1 - sum(self.citiesPartitions.values())
        ax4.pie([other] + list(self.citiesPartitions.values()),
                labels=["Другие"] + list(self.citiesPartitions.keys()), startangle=0)

        fig.tight_layout(pad=0.4, w_pad=0.5, h_pad=1.0)
        plt.savefig("graph.png")
        plt.show()


fileName = input("Введите название файла: ")
profession = input("Введите название профессии: ")
rep = Report(fileName, profession)
rep.print_file()
rep.generate_image()